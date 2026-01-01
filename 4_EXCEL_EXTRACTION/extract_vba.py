#!/usr/bin/env python3
"""
Extract all VBA macro code from Excel workbooks.
Outputs: individual .vba files per module, plus summary.md
"""

from openpyxl import load_workbook
from pathlib import Path
import re
from datetime import datetime

class VBAExtractor:
    def __init__(self, tracker_20_path: str, whole_school_path: str = None):
        """Load workbooks - note: openpyxl has limited VBA support
        We'll use python-pptx's approach via zipfile for full extraction"""
        self.tracker_20_path = tracker_20_path
        self.whole_school_path = whole_school_path
        self.output_dir = Path("vba_extractions")
        self.output_dir.mkdir(exist_ok=True)
        print(f"✓ VBA Extractor initialized. Output: {self.output_dir.absolute()}")
    
    def extract_vba_from_xlsm(self, xlsm_path: str, workbook_name: str = "Tracker 2.0"):
        """Extract VBA code from .xlsm file using zipfile"""
        import zipfile
        import xml.etree.ElementTree as ET
        
        print(f"\nExtracting VBA from: {workbook_name}")
        
        try:
            with zipfile.ZipFile(xlsm_path, 'r') as zip_ref:
                # List all files in the archive
                all_files = zip_ref.namelist()
                vba_files = [f for f in all_files if 'vbaProject' in f or 'VBA' in f]
                
                if not vba_files:
                    print(f"  ✗ No VBA files found in archive")
                    return []
                
                print(f"  Found {len(vba_files)} VBA-related files")
                
                # Extract VBA code files (usually in xl/vbaProject.bin or similar)
                vba_modules = {}
                
                # Try to find and extract vbaProject.bin
                vba_bin_path = 'xl/vbaProject.bin'
                if vba_bin_path in all_files:
                    print(f"  ✓ Found vbaProject.bin")
                    vba_data = zip_ref.read(vba_bin_path)
                    
                    # Extract readable text from binary (crude but effective)
                    vba_text = self._extract_text_from_binary(vba_data)
                    
                    # Save raw binary extraction
                    bin_output = self.output_dir / f"{workbook_name}_vbaProject.bin.txt"
                    bin_output.write_text(vba_text)
                    print(f"  ✓ Saved raw binary: {bin_output.name}")
                
                # Try to extract _rels/.rels and other XML files
                for file_path in all_files:
                    if file_path.endswith('.xml'):
                        try:
                            content = zip_ref.read(file_path).decode('utf-8', errors='ignore')
                            if any(vba_keyword in content for vba_keyword in ['VBA', 'Macro', 'Module', 'Function', 'Sub']):
                                xml_output = self.output_dir / f"{workbook_name}_{Path(file_path).name}.xml"
                                xml_output.write_text(content)
                                print(f"  ✓ Extracted: {xml_output.name}")
                        except:
                            pass
                
                return vba_files
        
        except Exception as e:
            print(f"  ✗ Error: {e}")
            return []
    
    def _extract_text_from_binary(self, data: bytes, min_length: int = 4) -> str:
        """Extract readable ASCII/UTF-8 strings from binary data"""
        result = []
        current = []
        
        for byte in data:
            if 32 <= byte <= 126:  # Printable ASCII
                current.append(chr(byte))
            else:
                if current and len(''.join(current)) >= min_length:
                    result.append(''.join(current))
                current = []
        
        if current and len(''.join(current)) >= min_length:
            result.append(''.join(current))
        
        return '\n'.join(result)
    
    def extract_sheet_validations(self, xlsm_path: str, workbook_name: str = "Tracker 2.0"):
        """Extract data validation rules from sheets (which define dropdowns, etc)"""
        print(f"\nExtracting Data Validations from: {workbook_name}")
        
        try:
            wb = load_workbook(xlsm_path)
            
            all_validations = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                if ws.data_validations:
                    sheet_validations = []
                    
                    for dv in ws.data_validations.dataValidation:
                        validation_info = {
                            "type": getattr(dv, 'type', None),
                            "formula1": getattr(dv, 'formula1', None),
                            "formula2": getattr(dv, 'formula2', None),
                            "showDropDown": getattr(dv, 'showDropDown', None),
                            "sqref": list(dv.sqref) if hasattr(dv, 'sqref') else [],
                            "showInputMessage": getattr(dv, 'showInputMessage', None),
                            "showErrorMessage": getattr(dv, 'showErrorMessage', None),
                            "errorTitle": getattr(dv, 'errorTitle', None),
                            "error": getattr(dv, 'error', None),
                            "promptTitle": getattr(dv, 'promptTitle', None),
                            "prompt": getattr(dv, 'prompt', None),
                        }
                        sheet_validations.append(validation_info)
                    
                    if sheet_validations:
                        all_validations[sheet_name] = sheet_validations
                        print(f"  ✓ {sheet_name}: {len(sheet_validations)} validation rules")
            
            if all_validations:
                output_file = self.output_dir / f"{workbook_name}_DataValidations.txt"
                self._write_validations_report(all_validations, output_file)
                print(f"  ✓ Saved: {output_file.name}")
            else:
                print(f"  - No data validations found")
            
            return all_validations
        
        except Exception as e:
            print(f"  - Skipped (openpyxl limitation): {type(e).__name__}")
            return {}
    
    def _write_validations_report(self, validations: dict, output_file: Path):
        """Write data validations to readable format"""
        lines = []
        lines.append("=" * 80)
        lines.append("DATA VALIDATION RULES")
        lines.append("=" * 80)
        lines.append(f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        for sheet_name, rules in validations.items():
            lines.append(f"\n{'SHEET: ' + sheet_name}")
            lines.append("-" * 80)
            
            for idx, rule in enumerate(rules, 1):
                lines.append(f"\nRule #{idx}")
                lines.append(f"  Type: {rule['type']}")
                lines.append(f"  Allow: {rule['allow']}")
                lines.append(f"  Cells: {', '.join(rule['sqref'])}")
                
                if rule['formula1']:
                    lines.append(f"  Formula 1: {rule['formula1']}")
                if rule['formula2']:
                    lines.append(f"  Formula 2: {rule['formula2']}")
                
                lines.append(f"  Show Dropdown: {rule['showDropDown']}")
                
                if rule['showInputMessage']:
                    lines.append(f"  Input Title: {rule['promptTitle']}")
                    lines.append(f"  Input Message: {rule['prompt']}")
                
                if rule['showErrorMessage']:
                    lines.append(f"  Error Title: {rule['errorTitle']}")
                    lines.append(f"  Error Message: {rule['error']}")
        
        output_file.write_text('\n'.join(lines))
    
    def extract_named_ranges(self, xlsm_path: str, workbook_name: str = "Tracker 2.0"):
        """Extract named ranges (used in formulas for readability)"""
        print(f"\nExtracting Named Ranges from: {workbook_name}")
        
        try:
            wb = load_workbook(xlsm_path)
            
            # Check if workbook has named_ranges or defined_names
            named_ranges = {}
            
            if hasattr(wb, 'named_ranges') and wb.named_ranges:
                for name, named_range in wb.named_ranges.items():
                    try:
                        named_ranges[name] = {
                            "name": name,
                            "targets": [(sheet, coord) for sheet, coord in named_range.destinations]
                        }
                    except:
                        pass
            elif hasattr(wb, 'defined_names') and wb.defined_names:
                for defined_name in wb.defined_names.definedName:
                    try:
                        named_ranges[defined_name.name] = {
                            "name": defined_name.name,
                            "value": defined_name.value
                        }
                    except:
                        pass
            else:
                print(f"  - No named ranges found")
                return {}
            
            if named_ranges:
                output_file = self.output_dir / f"{workbook_name}_NamedRanges.txt"
                lines = []
                lines.append("=" * 80)
                lines.append("NAMED RANGES")
                lines.append("=" * 80)
                lines.append(f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                
                for name, info in sorted(named_ranges.items()):
                    lines.append(f"\n{name}")
                    if 'targets' in info:
                        for sheet, coord in info['targets']:
                            lines.append(f"  → {sheet}!{coord}")
                    elif 'value' in info:
                        lines.append(f"  → {info['value']}")
                
                output_file.write_text('\n'.join(lines))
                print(f"  ✓ Found {len(named_ranges)} named ranges")
                print(f"  ✓ Saved: {output_file.name}")
                
                return named_ranges
            
            return {}
        
        except Exception as e:
            print(f"  - Skipped: {type(e).__name__}")
            return {}
    
    def extract_formulas_detailed(self, xlsm_path: str, workbook_name: str = "Tracker 2.0"):
        """Extract all formulas from workbook, organized by sheet and complexity"""
        print(f"\nExtracting Formulas from: {workbook_name}")
        
        try:
            wb = load_workbook(xlsm_path, data_only=False)
            
            all_formulas = {}
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_formulas = []
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            sheet_formulas.append({
                                "cell": cell.coordinate,
                                "formula": cell.value,
                                "length": len(cell.value),
                                "complexity": self._estimate_complexity(cell.value)
                            })
                
                if sheet_formulas:
                    all_formulas[sheet_name] = sheet_formulas
            
            if all_formulas:
                output_file = self.output_dir / f"{workbook_name}_Formulas_Detailed.txt"
                self._write_formulas_report(all_formulas, output_file)
                print(f"  ✓ Found {sum(len(f) for f in all_formulas.values())} formulas")
                print(f"  ✓ Saved: {output_file.name}")
                
                return all_formulas
            else:
                print(f"  - No formulas found")
            
            return {}
        
        except Exception as e:
            print(f"  ✗ Error: {e}")
            return {}
    
    def _estimate_complexity(self, formula: str) -> str:
        """Estimate formula complexity"""
        if formula.count('IF') == 0 and formula.count('INDEX') == 0:
            return "Simple"
        elif formula.count('IF') <= 2 and formula.count('INDEX') <= 1:
            return "Moderate"
        else:
            return "Complex"
    
    def _write_formulas_report(self, formulas: dict, output_file: Path):
        """Write formulas to organized report"""
        lines = []
        lines.append("=" * 100)
        lines.append("FORMULA ANALYSIS")
        lines.append("=" * 100)
        lines.append(f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        total_formulas = sum(len(f) for f in formulas.values())
        lines.append(f"Total Formulas: {total_formulas}\n")
        
        for sheet_name in sorted(formulas.keys()):
            sheet_formulas = formulas[sheet_name]
            
            # Sort by complexity
            simple = [f for f in sheet_formulas if f['complexity'] == 'Simple']
            moderate = [f for f in sheet_formulas if f['complexity'] == 'Moderate']
            complex_f = [f for f in sheet_formulas if f['complexity'] == 'Complex']
            
            lines.append(f"\n{'SHEET: ' + sheet_name}")
            lines.append("=" * 100)
            lines.append(f"Total: {len(sheet_formulas)} | Simple: {len(simple)} | Moderate: {len(moderate)} | Complex: {len(complex_f)}\n")
            
            # Complex formulas (most interesting)
            if complex_f:
                lines.append(f"\n--- COMPLEX FORMULAS ({len(complex_f)}) ---")
                for f in sorted(complex_f, key=lambda x: -x['length'])[:20]:  # Top 20 longest
                    lines.append(f"\n{f['cell']}: ({f['complexity']}, {f['length']} chars)")
                    # Format for readability
                    formatted = self._format_formula(f['formula'])
                    lines.append(formatted)
            
            # Moderate formulas
            if moderate:
                lines.append(f"\n--- MODERATE FORMULAS ({len(moderate)}) ---")
                for f in sorted(moderate, key=lambda x: -x['length'])[:10]:
                    lines.append(f"\n{f['cell']}: ({f['complexity']}, {f['length']} chars)")
                    formatted = self._format_formula(f['formula'])
                    lines.append(formatted)
            
            # Simple formulas (sample)
            if simple:
                lines.append(f"\n--- SIMPLE FORMULAS ({len(simple)} total) ---")
                for f in sorted(simple, key=lambda x: -x['length'])[:5]:
                    lines.append(f"{f['cell']}: {f['formula']}")
        
        output_file.write_text('\n'.join(lines))
    
    def _format_formula(self, formula: str) -> str:
        """Format formula for readability"""
        # Add newlines at IF statements
        formatted = formula.replace('IF(', '\n  IF(')
        formatted = formatted.replace('INDEX(', '\n    INDEX(')
        return "  " + formatted
    
    def create_summary_report(self):
        """Create summary of all extractions"""
        lines = []
        lines.append("=" * 100)
        lines.append("VBA EXTRACTION SUMMARY")
        lines.append("=" * 100)
        lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        lines.append(f"Output Directory: {self.output_dir.absolute()}\n")
        
        extracted_files = list(self.output_dir.glob("*"))
        lines.append(f"Files Generated: {len(extracted_files)}\n")
        
        for file_path in sorted(extracted_files):
            size_kb = file_path.stat().st_size / 1024
            lines.append(f"  - {file_path.name} ({size_kb:.1f} KB)")
        
        summary_file = self.output_dir / "00_EXTRACTION_SUMMARY.txt"
        summary_file.write_text('\n'.join(lines))
        print(f"\n✓ Summary saved: {summary_file.name}")

def main():
    tracker_path = r"C:\Users\robke\OneDrive\Desktop\Rob's FMS Scorecard\Rob's PE Movement Assessment Tracker 2.0.xlsm"
    
    print("=" * 100)
    print("VBA & MACRO CODE EXTRACTOR")
    print("=" * 100)
    
    extractor = VBAExtractor(tracker_path)
    
    # Extract all types of information
    extractor.extract_vba_from_xlsm(tracker_path, "Tracker_2.0")
    extractor.extract_sheet_validations(tracker_path, "Tracker_2.0")
    extractor.extract_named_ranges(tracker_path, "Tracker_2.0")
    extractor.extract_formulas_detailed(tracker_path, "Tracker_2.0")
    extractor.create_summary_report()
    
    print("\n" + "=" * 100)
    print("✓ VBA Extraction Complete!")
    print(f"✓ Check {extractor.output_dir.absolute()} for all extracted files")
    print("=" * 100)

if __name__ == "__main__":
    main()
