#!/usr/bin/env python3
"""
Comprehensive Excel extraction script.
Extracts ALL data/formulas from assessment sheets to .md files for manual AI review.
Does NOT attempt intelligent parsing—just dumps everything clearly.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from pathlib import Path
from datetime import datetime

class SimpleSkillExtractor:
    def __init__(self, tracker_20_path: str):
        """Load workbook with both values and formulas"""
        self.wb_data = load_workbook(tracker_20_path, data_only=True)  # Calculated values
        self.wb_formulas = load_workbook(tracker_20_path, data_only=False)  # Raw formulas
        self.output_dir = Path("skill_extractions")
        self.output_dir.mkdir(exist_ok=True)
        print(f"✓ Loaded workbook. Output: {self.output_dir.absolute()}")
    
    def extract_all_sheets(self):
        """Extract all non-trivial sheets to .md files"""
        skip_sheets = {"class list", "dashboard", "sheet", "whole school"}
        
        for sheet_name in self.wb_data.sheetnames:
            if not any(skip in sheet_name.lower() for skip in skip_sheets):
                print(f"\nProcessing: {sheet_name}")
                self._extract_sheet_to_md(sheet_name)
    
    def _extract_sheet_to_md(self, sheet_name: str):
        """Extract single sheet: all data, formulas, structure"""
        ws_data = self.wb_data[sheet_name]
        ws_formulas = self.wb_formulas[sheet_name]
        
        md_lines = []
        md_lines.append(f"# {sheet_name}\n")
        md_lines.append(f"**Extracted:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # 1. DIMENSIONS
        md_lines.append(f"\n## Sheet Dimensions\n")
        md_lines.append(f"- Max Row: {ws_data.max_row}\n")
        md_lines.append(f"- Max Column: {ws_data.max_column}\n")
        md_lines.append(f"- Dimensions: {ws_data.dimensions}\n")
        
        # 2. ALL CELLS WITH VALUES (GRID VIEW)
        md_lines.append(f"\n## Cell Grid (Values)\n")
        md_lines.append(self._generate_grid_view(ws_data))
        
        # 3. ALL CELLS WITH FORMULAS
        md_lines.append(f"\n## Cell Grid (Formulas)\n")
        md_lines.append(self._generate_formula_view(ws_formulas))
        
        # 4. MERGED CELLS
        merged_cells = list(ws_data.merged_cells.ranges)
        if merged_cells:
            md_lines.append(f"\n## Merged Cells\n")
            for merged_range in merged_cells:
                md_lines.append(f"- `{merged_range}`\n")
        
        # 5. CONDITIONAL FORMATTING RULES (if any)
        md_lines.append(f"\n## Conditional Formatting Rules\n")
        if ws_data.conditional_formatting:
            for rule_range, rules in ws_data.conditional_formatting._cf_rules.items():
                md_lines.append(f"- Range: `{rule_range}`\n")
                for rule in rules:
                    md_lines.append(f"  - Type: {rule.type}\n")
                    md_lines.append(f"  - Formula/Condition: {rule.formula if hasattr(rule, 'formula') else 'N/A'}\n")
        else:
            md_lines.append("*(No conditional formatting detected)*\n")
        
        # 6. CELL STYLES & COLORS
        md_lines.append(f"\n## Cell Formatting (Sample)\n")
        md_lines.append(self._extract_formatting_info(ws_data))
        
        # 7. NOTES/COMMENTS
        md_lines.append(f"\n## Cell Comments\n")
        comments_found = False
        for row in ws_data.iter_rows():
            for cell in row:
                if cell.comment:
                    md_lines.append(f"- `{cell.coordinate}`: {cell.comment.text}\n")
                    comments_found = True
        if not comments_found:
            md_lines.append("*(No comments found)*\n")
        
        # 8. RAW DATA BY ROW (for easy scanning)
        md_lines.append(f"\n## Raw Data by Row\n")
        md_lines.append("```\n")
        for row_idx in range(1, min(ws_data.max_row + 1, 100)):  # Limit to first 100 rows
            row_data = []
            for col_idx in range(1, ws_data.max_column + 1):
                cell = ws_data.cell(row_idx, col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(f"[{get_column_letter(col_idx)}{row_idx}={val}]")
            if row_data:
                md_lines.append(f"Row {row_idx}: {' '.join(row_data)}\n")
        md_lines.append("```\n")
        
        # Write file
        output_file = self.output_dir / f"{sheet_name}.md"
        output_file.write_text("".join(md_lines))
        print(f"  ✓ Saved: {output_file.name}")
    
    def _generate_grid_view(self, ws) -> str:
        """Generate table view of all cells (values)"""
        lines = []
        lines.append("| Row | " + " | ".join([f"Col {get_column_letter(c)}" for c in range(1, min(ws.max_column + 1, 15))]) + " |\n")
        lines.append("|-----|" + "|".join(["---"] * min(ws.max_column, 14)) + "|\n")
        
        for row_idx in range(1, min(ws.max_row + 1, 50)):  # Limit to first 50 rows for readability
            row_data = [str(row_idx)]
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                if val is None:
                    row_data.append("")
                else:
                    # Truncate long values
                    val_str = str(val)
                    if len(val_str) > 20:
                        val_str = val_str[:17] + "..."
                    row_data.append(val_str)
            lines.append("| " + " | ".join(row_data) + " |\n")
        
        return "".join(lines)
    
    def _generate_formula_view(self, ws) -> str:
        """Generate table view of formulas"""
        lines = []
        formulas_found = False
        
        lines.append("| Cell | Formula |\n")
        lines.append("|------|----------|\n")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    lines.append(f"| `{cell.coordinate}` | `{cell.value}` |\n")
                    formulas_found = True
        
        if not formulas_found:
            lines.append("| *(No formulas found)* | |\n")
        
        return "".join(lines)
    
    def _extract_formatting_info(self, ws) -> str:
        """Extract cell colors, fonts, alignment (sampling first 20 rows)"""
        lines = []
        lines.append("| Cell | Value | Font Color | Fill Color | Alignment |\n")
        lines.append("|------|-------|------------|------------|----------|\n")
        
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            for col_idx in range(1, min(ws.max_column + 1, 6)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value:
                    font_color = cell.font.color.rgb if cell.font and cell.font.color else "Default"
                    fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else "Default"
                    alignment = f"{cell.alignment.horizontal}/{cell.alignment.vertical}" if cell.alignment else "Default"
                    
                    lines.append(f"| `{cell.coordinate}` | {cell.value} | {font_color} | {fill_color} | {alignment} |\n")
        
        return "".join(lines)

if __name__ == "__main__":
    tracker_path = r"C:\Users\robke\OneDrive\Desktop\Rob's FMS Scorecard\Rob's PE Movement Assessment Tracker 2.0.xlsm"
    
    print("=" * 60)
    print("PE Assessment Skills Extractor")
    print("=" * 60)
    
    try:
        extractor = SimpleSkillExtractor(tracker_path)
        extractor.extract_all_sheets()
        print("\n" + "=" * 60)
        print("✓ Extraction complete!")
        print(f"✓ Check {extractor.output_dir.absolute()} for .md files")
        print("=" * 60)
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
